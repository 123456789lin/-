# coding:utf8
import os
import torch
import torch as t
from PIL import Image
from torch.utils import data
import numpy as np
from torchvision import transforms as T
import openpyxl

class ODIR(data.Dataset):

    def __init__(self, root,test=True,left=True):
        self.left = left

        self.root = root
        self.test = test
        #测试时的处理  读图片及id
        if self.test:
            imgt = [os.path.join(root + 'ODIR-5K_training/testing_set', img) for img in os.listdir(root + 'ODIR-5K_training/testing_set')]
            # imgt = [os.path.join(root + 'ODIR-5K_Testing_Images', img) for img in os.listdir(root + 'ODIR-5K_Testing_Images')]
            # imgt = [os.path.join(root + 'ODIR-5K_training/ODIR-5K_training', img) for img in os.listdir(root + 'ODIR-5K_training/ODIR-5K_training')]
            img_numt = len(imgt)
            imgt = sorted(imgt, key=lambda x: int(x.split('.')[-2].split('/')[-1].split('_')[-2]))

            list_1 = []
            list_2 = []
            for aaa in range(int(img_numt / 2)):
                if imgt[aaa * 2].split('.')[-2].split('/')[-1].split('_')[-1] == 'left':
                    list_1.append(imgt[aaa * 2])
                    list_1.append(imgt[aaa * 2 + 1])
                else:
                    list_1.append(imgt[aaa * 2 + 1])
                    list_1.append(imgt[aaa * 2])
                list_2.append(list_1.copy())
                list_1.clear()
            self.imgs = list_2
        else:
            #训练时的处理 读excel
            imgs = []
            total_list = []
            one_list = []
            two_list = []
            #read excel
            if self.left:
                # wb = openpyxl.load_workbook(root + 'ODIR-5K_training/trainingset_left_eye.xlsx')
                wb = openpyxl.load_workbook(root + 'ODIR-5K_training/left_eye.xlsx')
            else:
                # wb = openpyxl.load_workbook(root+'ODIR-5K_training/trainingset_right_eye.xlsx')
                wb = openpyxl.load_workbook(root + 'ODIR-5K_training/right_eye.xlsx')
            sheet1 = wb.get_sheet_by_name('Sheet1')
            #                        one_list           two_list...........
            # data structure:[[id,[left_pth,right_pth],[label*8]], ... ,[...]]
            for column in sheet1.rows:
                one_list.append(column[3].value)
                one_list.append(column[4].value)
                for i in range(7,15):
                    two_list.append(column[i].value)
             
                total_list.append(column[0].value)
                total_list.append(one_list.copy())
                total_list.append(two_list.copy())
            
                imgs.append(total_list.copy())
            
                one_list.clear()
                two_list.clear()
                total_list.clear()
            del imgs[0]
            self.imgs = imgs

        normalize = T.Normalize(mean=[0.485, 0.456, 0.406],std=[0.229, 0.224, 0.225])
        if self.test:
            self.transforms = T.Compose([
                T.Resize(224),
                T.CenterCrop(224),
                T.ToTensor(),
                normalize
                ])
        else:
            self.transforms = T.Compose([
                T.Resize(224),
                T.RandomResizedCrop(224),
                T.RandomHorizontalFlip(),
                T.ToTensor(),
                normalize
            ])

    def __getitem__(self, index):
        # picture
        information = self.imgs[index]
        if self.test:
            if self.left:
                left_img = information[0]
                left_data = Image.open(left_img)
                left_data = self.transforms(left_data)
                data = left_data
                id = left_img.split('.')[-2].split('/')[-1].split('_')[-2]
                label = id
            else:
                right_img = information[1]
                right_data = Image.open(right_img)
                right_data = self.transforms(right_data)
                data = right_data
                id = right_img.split('.')[-2].split('/')[-1].split('_')[-2]
                label = id
                # lable  -  存在着 id 的信息

        else:
            imgs_list = information[1]
            if self.left:
                left_img = imgs_list[0]
                # left_data = Image.open(self.root + 'ODIR-5K_training/training_set/' + left_img)
                left_data = Image.open(self.root + 'ODIR-5K_training/ODIR-5K_training/' + left_img)
                left_data = self.transforms(left_data)
                data = left_data

            else:
                right_img = imgs_list[1]
                # right_data = Image.open(self.root+'ODIR-5K_training/training_set/'+right_img)
                right_data = Image.open(self.root + 'ODIR-5K_training/ODIR-5K_training/' + right_img)
                right_data = self.transforms(right_data)
                data = right_data
            # lable
            label = np.array(information[2])
            label = t.from_numpy(label)
        return data,label

    def __len__(self):
        return len(self.imgs)
            

