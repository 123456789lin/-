import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment
# 选择左右眼,左眼left_eye = True,右眼为False
left_eye = True
training = True
# 输入文件路径,以/结尾
root = '/home/aaa/PycharmProjects/fgd/'

if left_eye:
    excel_row = 5
else:
    excel_row = 6

total_list = []
one_list = []
dic = {7:'正常眼底',8:'糖尿病',9:'青光眼',10:'白内障',11:'黄斑病',12:'高血压',13:'近视'}
# read excel
if training:
    wb = openpyxl.load_workbook(root + 'ODIR-5K_training/trainingset.xlsx')
    # wb = openpyxl.load_workbook(root + 'ODIR-5K_training/ODIR-5K_training-Chinese.xlsx')
else:
    wb = openpyxl.load_workbook(root + 'ODIR-5K_training/testingset.xlsx')
sheet1 = wb.get_sheet_by_name('Sheet1')
#                        one_list           two_list...........
# data structure:[[id,[left_pth,right_pth],[label*8]], ... ,[...]]
for column in sheet1.rows:
    # 前七列信息
    for i in range(7):
        if left_eye:
            if i == 4 or i == 6:
                one_list.append('None')
            else:
                one_list.append(column[i].value)
        else:
            if i == 3 or i == 5:
                one_list.append('None')
            else:
                one_list.append(column[i].value)
    # 由八类疾病的标签对应的疾病名称是否在关键字里边来区分左右眼
    label_1_num = 0
    keyword_num = 1
    # 前七类前七个标签数为1的数目
    for i in range(7,14):
        if dic[i] in one_list[excel_row]:
            one_list.append(1)
            # 统计前七个标签数为1的数目
            label_1_num += 1
        else:
            one_list.append(0)

    # 第八类, 第一种情况：关键词数目与前七个标签数为1的数目一致,则一定不含有其他疾病.(除去关键词，镜头污点,图片质量差)
    #        第二种情况：关键词数目与前七个标签数为1的数目不一致的情况,此种情况一定含有其他疾病.(除去关键词，镜头污点,图片质量差)
    # 统计关键词数目
    if ',' in one_list[excel_row]:
        keyword_num = len(one_list[excel_row].split(','))
        # 镜头污点，图片质量差，不能作为判断其他类的依据，去掉
        if '镜头污点' in one_list[excel_row]:
            keyword_num -= 1
        if '图片质量差' in one_list[excel_row]:
            keyword_num -= 1
        # 去除标注瑕疵：一个关键词重复出现，如id:3256 有不少这种情况
        defect = one_list[excel_row].split(',')
        defect_num = 0
        if len(defect)>1:
            for aaa in range(len(defect)-1):
                if defect[aaa] in defect[aaa+1]: defect_num += 1
            keyword_num = keyword_num - defect_num


    if keyword_num == label_1_num:
        one_list.append(0)
    else:
        one_list.append(1)

    total_list.append(one_list.copy())
    one_list.clear()
total_list[0] = ['编号', '病人年龄', '病人性别', '左眼眼底图像', '右眼眼底图像', '左眼诊断关键词', '右眼诊断关键词',
                 'N', 'D', 'G', 'C', 'A', 'H', 'M', 'O']
# 写入excel
wbe = Workbook()
wse = wbe.active
wse.title = "Sheet1"
# wse.create_sheet("Sheet0", 0)
# 调整列宽
if left_eye:
    wse.column_dimensions['D'].width = 12
    wse.column_dimensions['F'].width = 30
else:
    wse.column_dimensions['E'].width = 12
    wse.column_dimensions['G'].width = 30
# 写入信息
for j in total_list:
    wse.append(j)

# 设置对齐方式
# alignment=Alignment(horizontal='center', vertical='center')#水平'center', 'centerContinuous', 'justify', 'fill', 'general', 'distributed', 'left', 'right'
# for row in ['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O']:
#     for col in range(1,len(total_list)):
#         cc = row+str(col)
#         wse[cc].alignment = alignment

if training:
    if left_eye:
        wbe.save(root + 'ODIR-5K_training/trainingset_left_eye.xlsx')
        # wbe.save(root + 'ODIR-5K_training/left_eye.xlsx')
    else:
        wbe.save(root + 'ODIR-5K_training/trainingset_right_eye.xlsx')
        # wbe.save(root + 'ODIR-5K_training/right_eye.xlsx')
else:
    if left_eye:
        wbe.save(root+'ODIR-5K_training/testingset_left_eye.xlsx')
    else:
        wbe.save(root + 'ODIR-5K_training/testingset_right_eye.xlsx')





