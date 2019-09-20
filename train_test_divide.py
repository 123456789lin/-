import openpyxl
import random
from openpyxl import Workbook
from openpyxl.styles import Alignment
import os
import shutil

root = '/home/aaa/PycharmProjects/fgd/'
# 输入文件路径,以/结尾

root_excel = root+'ODIR-5K_training/ODIR-5K_training-Chinese.xlsx'

two_list = []
one_list = []

#read excel
wb = openpyxl.load_workbook(root_excel)
sheet1 = wb.get_sheet_by_name('Sheet1')
for column in sheet1.rows:
    for i in range(15):
        one_list.append(column[i].value)
    two_list.append(one_list.copy())
    one_list.clear()
title = two_list[0]
del two_list[0]
random.shuffle(two_list)
train_set = two_list[:2500]
test_set = two_list[2500:]
#按id排序
train_set = sorted(train_set, key=lambda x: int(x[0]))
test_set = sorted(test_set, key=lambda x: int(x[0]))
#创建训练集测试集文件夹，并复制相应的图片到对应的文件夹
#训练集
os.makedirs(root+'ODIR-5K_training/training_set')
for i in train_set:
    srcfile_left = root+'ODIR-5K_training/ODIR-5K_training/'+i[3]
    srcfile_right = root+'ODIR-5K_training/ODIR-5K_training/'+i[4]
    newfile_left = root+'ODIR-5K_training/training_set/'+i[3]
    newfile_right = root + 'ODIR-5K_training/training_set/'+i[4]
    shutil.copyfile(srcfile_left,newfile_left)
    shutil.copyfile(srcfile_right,newfile_right)
#测试集
os.makedirs(root+'ODIR-5K_training/testing_set')
for i in test_set:
    srcfile_left = root+'ODIR-5K_training/ODIR-5K_training/' +i[3]
    srcfile_right = root+'ODIR-5K_training/ODIR-5K_training/' +i[4]
    newfile_left = root+'ODIR-5K_training/testing_set/'+i[3]
    newfile_right = root + 'ODIR-5K_training/testing_set/'+i[4]
    shutil.copyfile(srcfile_left,newfile_left)
    shutil.copyfile(srcfile_right,newfile_right)

# 写入excel 训练集
wbe = Workbook()
wse = wbe.active
wse.title = "Sheet1"
# wse.create_sheet("Sheet0", 0)
# 调整列宽
wse.column_dimensions['E'].width = 12
wse.column_dimensions['G'].width = 30
wse.column_dimensions['D'].width = 12
wse.column_dimensions['F'].width = 30
# 写入信息
wse.append(title)
for j in train_set:
    wse.append(j)
# 设置对齐方式
# alignment=Alignment(horizontal='center', vertical='center')#水平'center', 'centerContinuous', 'justify', 'fill', 'general', 'distributed', 'left', 'right'
# for row in ['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O']:
#     for col in range(1,len(train_set)):
#         cc = row+str(col)
#         wse[cc].alignment = alignment
wbe.save(root + 'ODIR-5K_training/trainingset.xlsx')


# 写入excel  测试集
wbe = Workbook()
wse = wbe.active
wse.title = "Sheet1"
# wse.create_sheet("Sheet0", 0)
# 调整列宽
wse.column_dimensions['E'].width = 15
wse.column_dimensions['G'].width = 30
wse.column_dimensions['D'].width = 15
wse.column_dimensions['F'].width = 30
# 写入信息
wse.append(title)
for j in test_set:
    wse.append(j)
# 设置对齐方式
# alignment=Alignment(horizontal='center', vertical='center')#水平'center', 'centerContinuous', 'justify', 'fill', 'general', 'distributed', 'left', 'right'
# for row in ['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O']:
#     for col in range(1,len(test_set)):
#         cc = row+str(col)
#         wse[cc].alignment = alignment
wbe.save(root + 'ODIR-5K_training/testingset.xlsx')